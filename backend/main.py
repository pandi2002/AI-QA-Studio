import json
import traceback
from fastapi.staticfiles import StaticFiles

from fastapi import (
    FastAPI,
    Form,
    File,
    UploadFile,
    Body,
)
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
)

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER 
from reportlab.lib.colors import HexColor 

from datetime import datetime

from routes.automation_route import router as automation_router

from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from routes.bug_report import router as bug_report_router
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.sql_service import generate_sql
from services.ai_provider import (
    generate_testcases,
    generate_playwright,
    generate_review,
)

app = FastAPI(
    title="AI QA Studio",
    version="2.0"
)
from pathlib import Path
from fastapi.responses import RedirectResponse

ALLURE_REPORT_DIR = Path(__file__).parent / "automation" / "allure-report"

# Create the directory if it doesn't exist
ALLURE_REPORT_DIR.mkdir(parents=True, exist_ok=True)

@app.get("/allure-report-url")
def get_allure_report_url():
    index_path = ALLURE_REPORT_DIR / "index.html"
    if index_path.exists() and index_path.stat().st_size > 0:
        return {"url": "/allure-report/index.html", "is_local": True}
    return {"url": "https://pandi2002.github.io/AI-QA-Studio/", "is_local": False}

app.mount(
    "/allure-report",
    StaticFiles(directory=ALLURE_REPORT_DIR),
    name="allure-report",
)

app.include_router(bug_report_router)
app.include_router(automation_router)


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def home():
    return {
        "message": "AI QA Studio Backend Running"
    }


# ======================================================
# Generate Test Cases
# ======================================================

@app.post("/generate-testcases")
async def generate_testcases_api(

    requirement: str = Form(""),

    testingTypes: str = Form("[]"),

    designTechniques: str = Form("[]"),

    outputOptions: str = Form("[]"),

    provider: str = Form("gemini"),

    images: list[UploadFile] = File(default=[]),

):

    try:

        testing_types = json.loads(testingTypes)

        design_techniques = json.loads(designTechniques)

        output_options = json.loads(outputOptions)

        result = await generate_testcases(

            provider=provider,

            requirement=requirement,

            testing_types=testing_types,

            design_techniques=design_techniques,

            images=images,

        )

        return {
            "result": result
        }

    except Exception as e:
        traceback.print_exc()   # Prints the full error in the terminal

        return {
            "result": str(e)    # Sends the error back to the frontend
        }


# ======================================================
# Generate Playwright
# ======================================================

@app.post("/generate-playwright")
async def generate_playwright_api(data: dict = Body(...)):

    try:

        result = await generate_playwright(provider=data["provider"],requirement=data["requirement"],testcase_data=data["testcase_data"],)

        return {
            "result": result
        }

    except Exception as e:

        traceback.print_exc()

        return {
            "result": str(e)
        }


# ======================================================
# Export Excel
# ======================================================

@app.post("/export-excel")
def export_excel(data: dict = Body(...)):

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Priority",
        "Scenario",
        "Preconditions",
        "Steps",
        "Test Data",
        "Expected Result",
        "Design Technique",
    ]

    fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid",
    )

    font = Font(
        bold=True,
        color="FFFFFF",
    )

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = fill
        cell.font = font

    # Data rows
    row = 2

    for tc in data.get("testCases", []):

        ws.cell(row=row, column=1).value = tc.get("testCaseId", "")
        ws.cell(row=row, column=2).value = tc.get("category", "")
        ws.cell(row=row, column=3).value = tc.get("priority", "")
        ws.cell(row=row, column=4).value = tc.get("scenario", "")

        ws.cell(row=row, column=5).value = "\n".join(
            tc.get("preconditions", [])
        )
        steps = tc.get("steps", [])
        formatted_steps = "\n\n".join(
             f"Step {index}: {step}"
             for index, step in enumerate(steps, start=1)
             )
        ws.cell(row=row, column=6).value = formatted_steps

        ws.cell(row=row, column=7).value = tc.get("testData", "")
        ws.cell(row=row, column=8).value = tc.get("expectedResult", "")
        ws.cell(row=row, column=9).value = tc.get("designTechnique", "")

        row += 1

    # Auto-size columns
    for column_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 5, 50)

        EXPORT_DIR = Path(__file__).parent / "exports"
        EXPORT_DIR.mkdir(exist_ok=True)

        file_path = EXPORT_DIR / "Generated_TestCases.xlsx"

        wb.save(file_path)
        return FileResponse(
            path=file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="Generated_TestCases.xlsx",

        )
# ======================================================
# Export PDF
# ======================================================

@app.post("/export-pdf")
def export_pdf(data: dict = Body(...)):

    EXPORT_DIR = Path(__file__).parent / "exports"
    EXPORT_DIR.mkdir(exist_ok=True)

    file_path = EXPORT_DIR / "Generated_TestCases.pdf"

    doc = SimpleDocTemplate(str(file_path))

    styles = getSampleStyleSheet()

    story = []

    # ---------------------------------------------
    # Title
    # ---------------------------------------------

    title = styles["Title"]
    title.alignment = TA_CENTER
    title.textColor = HexColor("#1F4E79")

    story.append(
        Paragraph("AI QA Studio", title)
    )

    story.append(
        Paragraph(
            "<b>Generated Test Cases Report</b>",
            styles["Heading2"],
        )
    )

    story.append(
        Paragraph(
            f"Generated On : {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",
            styles["Normal"],
        )
    )

    story.append(Spacer(1, 20))

    # ---------------------------------------------
    # Test Cases
    # ---------------------------------------------

    for tc in data.get("testCases", []):

        story.append(
            Paragraph(
                f"<b>{tc.get('testCaseId','')}</b>",
                styles["Heading2"],
            )
        )

        story.append(
            Paragraph(
                f"<b>Category :</b> {tc.get('category','')}",
                styles["BodyText"],
            )
        )

        story.append(
            Paragraph(
                f"<b>Priority :</b> {tc.get('priority','')}",
                styles["BodyText"],
            )
        )

        story.append(Spacer(1, 10))

        # Scenario

        story.append(
            Paragraph(
                "<b>Scenario</b>",
                styles["Heading3"],
            )
        )

        story.append(
            Paragraph(
                tc.get("scenario",""),
                styles["BodyText"],
            )
        )

        story.append(Spacer(1, 10))

        # Preconditions

        story.append(
            Paragraph(
                "<b>Preconditions</b>",
                styles["Heading3"],
            )
        )

        preconditions = tc.get("preconditions", [])

        for index, precondition in enumerate(preconditions, start=1):

            story.append(
                Paragraph(
                    f"Precondition {index}: {precondition}",
                    styles["BodyText"],
                )
            )

        story.append(Spacer(1, 10))

        # Steps

        story.append(
            Paragraph(
                "<b>Steps</b>",
                styles["Heading3"],
            )
        )

        steps = tc.get("steps", [])

        for index, step in enumerate(steps, start=1):

            story.append(
                Paragraph(
                    f"Step {index}: {step}",
                    styles["BodyText"],
                )
            )

        story.append(Spacer(1, 10))

        # Test Data

        story.append(
            Paragraph(
                "<b>Test Data</b>",
                styles["Heading3"],
            )
        )

        story.append(
            Paragraph(
                tc.get("testData",""),
                styles["BodyText"],
            )
        )

        story.append(Spacer(1, 10))

        # Expected Result

        story.append(
            Paragraph(
                "<b>Expected Result</b>",
                styles["Heading3"],
            )
        )

        story.append(
            Paragraph(
                tc.get("expectedResult",""),
                styles["BodyText"],
            )
        )

        story.append(Spacer(1, 10))

        # Design Technique

        story.append(
            Paragraph(
                "<b>Design Technique</b>",
                styles["Heading3"],
            )
        )

        story.append(
            Paragraph(
                tc.get("designTechnique",""),
                styles["BodyText"],
            )
        )

        story.append(Spacer(1, 25))

    # ---------------------------------------------
    # Build PDF
    # ---------------------------------------------

    doc.build(story)

    return FileResponse(
        path=file_path,
        media_type="application/pdf",
        filename="Generated_TestCases.pdf",
    )


# ======================================================
# AI Review
# ======================================================

@app.post("/generate-review")
async def generate_review_api(data: dict = Body(...)):

    try:

        result = await generate_review(provider=data["provider"],

            testcase_data=data["testcase_data"],)

        return {
            "result": result
        }

    except Exception:

        traceback.print_exc()

        return {
            "result": "Failed to generate AI Review."
        }

# ======================================================
# SQL Generator
# ======================================================

@app.post("/generate-sql")
async def generate_sql_api(data: dict = Body(...)):

    try:

        result = await generate_sql(
            provider=data["provider"],
            requirement=data["requirement"],
            testcase_data=data.get("testcase_data"),
        )

        return {
            "result": result
        }

    except Exception as e:

        traceback.print_exc()

        return {
            "result": str(e)
        }